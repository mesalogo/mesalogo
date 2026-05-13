"""
Auto-converted from Flask Blueprint → FastAPI APIRouter
Source files: utils.py, base.py, agents.py, environment.py, export.py, monitoring.py, rules.py
"""
import logging
from typing import Optional
from fastapi import APIRouter, HTTPException, Query, Request, Depends
from fastapi.responses import JSONResponse, StreamingResponse, Response
from core.config import settings
from core.dependencies import get_current_user, get_admin_user, filter_user_tasks, can_access_task

logger = logging.getLogger(__name__)

router = APIRouter()

# ============================================================
# Source: base.py
# ============================================================

"""
Action Tasks 基础管理 - CRUD
"""

from sqlalchemy.exc import IntegrityError

from app.models import (
    ActionTask, ActionSpace, ActionTaskAgent, ActionTaskEnvironmentVariable,
    Conversation, ConversationAgent, Message, Agent, AutonomousTask, AutonomousTaskExecution,
    ActionSpaceEnvironmentVariable, ActionSpaceObserver, ActionSpaceRole, PublishedTask,
    Role, RoleVariable, RuleSet, User, db,
    AgentVariable, Rule, RuleSetRule, ActionSpaceRuleSet, RuleTriggerLog
)
from app.services.action_task_service import ActionTaskService
from app.services.agent_variable_service import AgentVariableService
from app.services.workspace_service import workspace_service
from app.utils.uuid_utils import UUIDValidator

def validate_task_access(task_id, current_user):
    """验证用户对Action Task的访问权限"""
    task = ActionTask.query.get(task_id)
    if not task:
        return False, None, ({'success': False, 'message': '任务不存在'}, 404)
    if not can_access_task(task, current_user):
        return False, None, ({'success': False, 'message': '无权限访问此任务'}, 403)
    return True, task, None

import logging
logger = logging.getLogger(__name__)

# 创建Blueprint

@router.get('/action-tasks')
def get_action_tasks(request: Request, current_user=Depends(get_current_user)):
    """获取行动任务列表（根据用户权限过滤）

    优化：用 SQL 子查询替代 N+1 循环查询，将 ~30万次 SQL 降为 1 次。
    """
    from sqlalchemy import func, case
    from sqlalchemy.orm import aliased

    include_agents = request.query_params.get('include_agents', 'false').lower() == 'true'
    # 是否包含实验克隆任务（默认排除，前端之前在 JS 中过滤，现在后端直接过滤）
    include_experiment_clones = request.query_params.get('include_experiment_clones', 'false').lower() == 'true'
    # 分页参数（可选，不传则返回全部）
    page_str = request.query_params.get('page')
    limit_str = request.query_params.get('limit')

    # ── 子查询：各维度的 count ──
    agent_count_sq = (
        db.session.query(
            ActionTaskAgent.action_task_id,
            func.count(ActionTaskAgent.id).label('agent_count')
        )
        .group_by(ActionTaskAgent.action_task_id)
        .subquery('agent_cnt')
    )

    message_count_sq = (
        db.session.query(
            Message.action_task_id,
            func.count(Message.id).label('message_count')
        )
        .group_by(Message.action_task_id)
        .subquery('msg_cnt')
    )

    conversation_count_sq = (
        db.session.query(
            Conversation.action_task_id,
            func.count(Conversation.id).label('conversation_count')
        )
        .group_by(Conversation.action_task_id)
        .subquery('conv_cnt')
    )

    # 自主任务统计（通过 conversation JOIN autonomous_tasks）
    autonomous_sq = (
        db.session.query(
            Conversation.action_task_id,
            func.count(AutonomousTask.id).label('total_autonomous'),
            func.sum(case(
                (AutonomousTask.status == 'active', 1),
                else_=0
            )).label('active_autonomous')
        )
        .join(AutonomousTask, AutonomousTask.conversation_id == Conversation.id)
        .group_by(Conversation.action_task_id)
        .subquery('auto_cnt')
    )

    # 发布状态
    _pt = PublishedTask.__table__
    published_sq = (
        db.session.query(
            _pt.c.action_task_id,
            _pt.c.share_token,
            _pt.c.access_type,
            _pt.c.mode,
            _pt.c.view_count,
            _pt.c.created_at.label('pub_created_at')
        )
        .filter(_pt.c.is_active == True)
        .subquery('pub')
    )

    # ── 主查询：一次性 JOIN 获取所有数据 ──
    # 注意: User.display_name 是 @property（从 profile JSON 读取），不能用在 SQL 中
    base_query = (
        db.session.query(
            ActionTask,
            func.coalesce(agent_count_sq.c.agent_count, 0).label('agent_count'),
            func.coalesce(message_count_sq.c.message_count, 0).label('message_count'),
            func.coalesce(conversation_count_sq.c.conversation_count, 0).label('conversation_count'),
            func.coalesce(autonomous_sq.c.total_autonomous, 0).label('total_autonomous'),
            func.coalesce(autonomous_sq.c.active_autonomous, 0).label('active_autonomous'),
            ActionSpace.name.label('action_space_name'),
            User.username.label('creator_username'),
            User.profile.label('creator_profile'),
            published_sq.c.share_token,
            published_sq.c.access_type,
            published_sq.c.mode.label('pub_mode'),
            published_sq.c.view_count,
            published_sq.c.pub_created_at,
        )
        .outerjoin(agent_count_sq, agent_count_sq.c.action_task_id == ActionTask.id)
        .outerjoin(message_count_sq, message_count_sq.c.action_task_id == ActionTask.id)
        .outerjoin(conversation_count_sq, conversation_count_sq.c.action_task_id == ActionTask.id)
        .outerjoin(autonomous_sq, autonomous_sq.c.action_task_id == ActionTask.id)
        .outerjoin(ActionSpace, ActionSpace.id == ActionTask.action_space_id)
        .outerjoin(User, User.id == ActionTask.user_id)
        .outerjoin(published_sq, published_sq.c.action_task_id == ActionTask.id)
    )

    # 权限过滤
    if not current_user.is_admin:
        from sqlalchemy import or_
        base_query = base_query.filter(
            or_(
                ActionTask.user_id == current_user.id,
                ActionTask.is_shared == True
            )
        )

    # 过滤实验克隆任务（默认排除，将 33,597 → ~7 条）
    if not include_experiment_clones:
        base_query = base_query.filter(
            (ActionTask.is_experiment_clone == False) | (ActionTask.is_experiment_clone == None)
        )

    # 排序
    base_query = base_query.order_by(ActionTask.updated_at.desc())

    # 分页
    if page_str:
        page = int(page_str)
        limit = int(limit_str) if limit_str else 20
        total = base_query.count()
        rows = base_query.offset((page - 1) * limit).limit(limit).all()
    else:
        rows = base_query.all()
        total = len(rows)

    # ── 如果需要 agents 明细，批量预加载 ──
    agent_map = {}
    if include_agents:
        task_ids = [row.ActionTask.id for row in rows]
        if task_ids:
            agent_rows = (
                db.session.query(ActionTaskAgent, Agent)
                .join(Agent, Agent.id == ActionTaskAgent.agent_id)
                .filter(ActionTaskAgent.action_task_id.in_(task_ids))
                .all()
            )
            for ta, agent in agent_rows:
                agent_map.setdefault(ta.action_task_id, []).append({
                    'id': agent.id,
                    'name': agent.name,
                    'description': agent.description,
                    'avatar': agent.avatar,
                    'is_default': ta.is_default,
                    'is_observer': agent.is_observer,
                })

    # ── 组装结果 ──
    result = []
    for row in rows:
        task = row.ActionTask
        is_published = row.share_token is not None
        publish_info = None
        if is_published:
            publish_info = {
                'share_token': row.share_token,
                'access_type': row.access_type,
                'mode': row.pub_mode,
                'view_count': row.view_count,
                'created_at': row.pub_created_at.isoformat() if row.pub_created_at else None,
            }

        # User.display_name 是 @property，从 profile JSON 中读取
        creator_display_name = None
        if row.creator_profile and isinstance(row.creator_profile, dict):
            creator_display_name = row.creator_profile.get('display_name', row.creator_username)
        elif row.creator_username:
            creator_display_name = row.creator_username

        task_data = {
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'status': task.status,
            'mode': task.mode,
            'created_at': task.created_at.isoformat() if task.created_at else None,
            'updated_at': task.updated_at.isoformat() if task.updated_at else None,
            'agent_count': row.agent_count,
            'message_count': row.message_count,
            'conversation_count': row.conversation_count,
            'autonomous_task_count': row.total_autonomous,
            'total_autonomous_task_count': row.total_autonomous,
            'active_autonomous_task_count': row.active_autonomous,
            'action_space_id': task.action_space_id,
            'action_space_name': row.action_space_name,
            'is_published': is_published,
            'publish_info': publish_info,
            'is_experiment_clone': task.is_experiment_clone,
            'creator_display_name': creator_display_name,
        }

        if include_agents:
            task_data['agents'] = agent_map.get(task.id, [])

        result.append(task_data)

    response = {'action_tasks': result}
    # 分页时返回额外信息
    if page_str:
        page = int(page_str)
        limit = int(limit_str) if limit_str else 20
        response['total'] = total
        response['page'] = page
        response['limit'] = limit
        response['total_pages'] = (total + limit - 1) // limit
    return response

@router.get('/action-tasks/{task_id}')
def get_action_task(task_id, current_user=Depends(get_current_user)):
    """获取特定行动任务详情"""
    # 验证UUID格式
    validation_error = UUIDValidator.validate_request_uuid(task_id, "task_id")
    if validation_error:
        raise HTTPException(status_code=validation_error.get("code", 400), detail=validation_error)

    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 检查用户是否有权限访问此任务
    if not can_access_task(task, current_user):
        raise HTTPException(status_code=403, detail={'error': '无权限访问此任务'})

    # 获取行动任务的智能体
    task_agents = ActionTaskAgent.query.filter_by(action_task_id=task_id).all()
    agents = []

    for ta in task_agents:
        agent = Agent.query.get(ta.agent_id)
        if agent:
            # 获取智能体的角色信息
            agent_role = None
            if hasattr(agent, 'role_id') and agent.role_id:
                agent_role = Role.query.get(agent.role_id)

            # 构建包含角色名称的智能体名称
            role_name = agent_role.name if agent_role else "智能助手"

            agents.append({
                'id': agent.id,
                'name': agent.name,
                'role_name': role_name,
                'description': agent.description,
                'avatar': agent.avatar,
                'is_default': ta.is_default,
                'is_observer': agent.is_observer,  # 添加是否为监督者标记
            })

    # 获取行动空间信息
    action_space = ActionSpace.query.get(task.action_space_id) if task.action_space_id else None
    rule_set = RuleSet.query.get(task.rule_set_id) if task.rule_set_id else None

    result = {
        'id': task.id,
        'title': task.title,
        'description': task.description,
        'status': task.status,
        'mode': task.mode,
        'created_at': task.created_at.isoformat() if task.created_at else None,
        'updated_at': task.updated_at.isoformat() if task.updated_at else None,
        'agents': agents,
        'action_space_id': task.action_space_id,  # 添加action_space_id字段
        'action_space': {
            'id': action_space.id,
            'name': action_space.name,
            'description': action_space.description
        } if action_space else None,
        'rule_set_id': task.rule_set_id,  # 添加rule_set_id字段
        'rule_set': {
            'id': rule_set.id,
            'name': rule_set.name,
            'description': rule_set.description
        } if rule_set else None
    }

    return result

@router.post('/action-tasks')
async def create_action_task(request: Request, current_user=Depends(get_current_user)):
    """创建新行动任务"""
    # 获取当前用户
    data = await request.json()

    # 验证必填字段
    if not data.get('title'):
        raise HTTPException(status_code=400, detail={'error': '缺少必填字段: title'})

    # 配额检查
    from app.services.subscription_service import SubscriptionService
    quota_result = SubscriptionService.check_quota(current_user.id, 'tasks')
    if not quota_result['allowed']:
        raise HTTPException(status_code=403, detail={
            'error': '已达到计划限额',
            'message': f'您的计划最多可创建 {quota_result["limit"]} 个行动任务',
            'quota': quota_result
        })

    # 设置多租户字段
    # ActionTask 使用 user_id 而不是 created_by
    user_id = None
    is_shared = False

    if current_user:
        # 所有用户创建的任务都记录创建者
        user_id = current_user.id
        is_shared = data.get('is_shared', False)  # 默认私有，可勾选共享

    # 创建新行动任务
    new_task = ActionTask(
        title=data.get('title'),
        description=data.get('description'),
        mode=data.get('mode', 'sequential'),
        status='active',
        rule_set_id=data.get('rule_set_id'),
        action_space_id=data.get('action_space_id'),
        user_id=user_id,
        is_shared=is_shared
    )

    # 保存行动任务
    db.session.add(new_task)
    db.session.flush()  # 获取ID但不提交事务

    # 如果有关联的行动空间，从行动空间初始化环境变量
    if new_task.action_space_id:
        action_space = ActionSpace.query.get(new_task.action_space_id)
        if action_space:
            try:
                # 检查请求中是否已经包含了环境变量数据
                has_space_vars_in_request = False
                if data.get('environment_variables'):
                    # 查看是否有标记为space来源的环境变量
                    for var_data in data.get('environment_variables'):
                        if var_data.get('source') == 'space':
                            has_space_vars_in_request = True
                            break

                # 只有在请求中没有行动空间环境变量时才从行动空间获取
                if not has_space_vars_in_request:
                    # 从行动空间获取传统环境变量定义
                    space_env_vars = ActionSpaceEnvironmentVariable.query.filter_by(
                        action_space_id=new_task.action_space_id
                    ).all()

                    # 为每个传统行动空间环境变量创建对应的任务环境变量
                    for space_var in space_env_vars:
                        env_var = ActionTaskEnvironmentVariable(
                            name=space_var.name,
                            label=space_var.label,
                            value=space_var.value,
                            action_task_id=new_task.id
                        )
                        db.session.add(env_var)

                    # 获取行动空间绑定的共享环境变量
                    from app.models import ActionSpaceSharedVariable, SharedEnvironmentVariable
                    shared_bindings = db.session.query(ActionSpaceSharedVariable, SharedEnvironmentVariable).join(
                        SharedEnvironmentVariable, ActionSpaceSharedVariable.shared_variable_id == SharedEnvironmentVariable.id
                    ).filter(ActionSpaceSharedVariable.action_space_id == new_task.action_space_id).all()

                    # 为每个共享环境变量创建对应的任务环境变量
                    for binding, shared_var in shared_bindings:
                        env_var = ActionTaskEnvironmentVariable(
                            name=shared_var.name,
                            label=shared_var.label,
                            value=shared_var.value,
                            shared_variable_id=shared_var.id,  # 关联共享变量ID
                            is_readonly=shared_var.is_readonly,  # 继承只读属性
                            action_task_id=new_task.id
                        )
                        db.session.add(env_var)

                # 如果请求中包含环境变量，创建它们
                if data.get('environment_variables'):
                    for var_data in data.get('environment_variables'):
                        env_var = ActionTaskEnvironmentVariable(
                            name=var_data.get('name'),
                            label=var_data.get('label', var_data.get('name', '').replace('_', ' ').title()),
                            value=var_data.get('value'),
                            type=var_data.get('type', 'text'),
                            action_task_id=new_task.id
                        )
                        db.session.add(env_var)

                # 环境变量初始化部分不再创建监督者智能体，移到单独的部分
            except Exception as e:
                logger.error(f"初始化环境变量或监督者失败: {str(e)}")
                # 继续处理，不中断任务创建

    # 从行动空间角色创建参与智能体
    if new_task.action_space_id:
        action_space = ActionSpace.query.get(new_task.action_space_id)
        if action_space:
            try:
                # 获取行动空间中的所有普通角色
                space_roles = ActionSpaceRole.query.filter_by(
                    action_space_id=new_task.action_space_id
                ).all()

                # 为每个普通角色创建智能体实例
                participant_count = 0
                for space_role in space_roles:
                    role = Role.query.get(space_role.role_id)
                    if role:
                        # 创建智能体实例
                        agent = Agent(
                            name=role.name,
                            description=role.description,
                            avatar=role.avatar,
                            settings=role.settings,
                            action_task_id=new_task.id,
                            role_id=role.id,
                            type='agent',  # 标记为普通智能体类型
                            is_observer=False,  # 标记为非监督者
                            additional_prompt=space_role.additional_prompt  # 设置额外提示词
                        )
                        db.session.add(agent)
                        db.session.flush()  # 获取agent.id

                        participant_count += 1

                        # 第一个参与者智能体设为默认
                        is_default = (participant_count == 1)

                        # 创建行动任务-智能体关联
                        task_agent = ActionTaskAgent(
                            action_task_id=new_task.id,
                            agent_id=agent.id,
                            is_default=is_default
                        )
                        db.session.add(task_agent)

                        # 查找该角色在当前行动空间的环境变量
                        role_vars = RoleVariable.query.filter_by(
                            role_id=role.id,
                            action_space_id=new_task.action_space_id
                        ).all()

                        # 为智能体创建这些环境变量
                        for role_var in role_vars:
                            try:
                                # 将角色变量转换为智能体变量
                                AgentVariableService.create_variable(
                                    agent_id=agent.id,
                                    name=role_var.name,
                                    value=role_var.value,
                                    is_public=True,
                                    label=role_var.label
                                )
                                logger.info(f"已为智能体 {agent.id} 创建环境变量: {role_var.name}")
                            except Exception as e:
                                logger.error(f"为智能体 {agent.id} 创建环境变量 {role_var.name} 失败: {str(e)}")

                logger.info(f"已从行动空间角色创建参与智能体")
            except Exception as e:
                logger.error(f"从行动空间角色创建参与智能体失败: {str(e)}")
                # 继续处理，不中断任务创建

            # 从行动空间监督者角色创建监督者智能体
            try:
                if data.get('include_observers', True):  # 默认包含监督者
                    # 获取行动空间中的所有监督者角色
                    space_observers = ActionSpaceObserver.query.filter_by(
                        action_space_id=new_task.action_space_id
                    ).all()

                    # 为每个监督者角色创建智能体实例
                    observer_count = 0
                    for space_observer in space_observers:
                        role = Role.query.get(space_observer.role_id)
                        if role:
                            # 创建监督者智能体实例
                            observer_agent = Agent(
                                name=role.name,
                                description=role.description,
                                avatar=role.avatar,
                                settings=role.settings,
                                action_task_id=new_task.id,
                                role_id=role.id,
                                type='observer',  # 标记为监督者类型
                                is_observer=True,  # 标记为监督者
                                additional_prompt=space_observer.additional_prompt  # 设置额外提示词
                            )
                            db.session.add(observer_agent)
                            db.session.flush()  # 获取observer_agent.id

                            observer_count += 1

                            # 创建行动任务-智能体关联（监督者不设为默认智能体）
                            task_agent = ActionTaskAgent(
                                action_task_id=new_task.id,
                                agent_id=observer_agent.id,
                                is_default=False
                            )
                            db.session.add(task_agent)

                            # 查找该角色在当前行动空间的环境变量
                            role_vars = RoleVariable.query.filter_by(
                                role_id=role.id,
                                action_space_id=new_task.action_space_id
                            ).all()

                            # 为监督者智能体创建这些环境变量
                            for role_var in role_vars:
                                try:
                                    # 将角色变量转换为智能体变量
                                    AgentVariableService.create_variable(
                                        agent_id=observer_agent.id,
                                        name=role_var.name,
                                        value=role_var.value,
                                        is_public=True,
                                        label=role_var.label
                                    )
                                    logger.info(f"已为监督者智能体 {observer_agent.id} 创建环境变量: {role_var.name}")
                                except Exception as e:
                                    logger.error(f"为监督者智能体 {observer_agent.id} 创建环境变量 {role_var.name} 失败: {str(e)}")

                    logger.info(f"已从行动空间监督者角色创建监督者智能体，共 {observer_count} 个")
            except Exception as e:
                logger.error(f"从行动空间监督者角色创建监督者智能体失败: {str(e)}")
                # 继续处理，不中断任务创建

    # 兼容旧版本：如果提供了agent_ids，也添加这些智能体
    if data.get('agent_ids'):
        for idx, agent_id in enumerate(data.get('agent_ids')):
            # 获取智能体并设置 action_task_id
            agent = Agent.query.get(agent_id)
            if agent:
                agent.action_task_id = new_task.id

                # 获取行动空间中角色的额外提示词
                if agent.role_id and new_task.action_space_id:
                    action_space_role = ActionSpaceRole.query.filter_by(
                        action_space_id=new_task.action_space_id,
                        role_id=agent.role_id
                    ).first()

                    if action_space_role and action_space_role.additional_prompt:
                        agent.additional_prompt = action_space_role.additional_prompt

                # 获取智能体对应的角色
                if agent.role_id and new_task.action_space_id:
                    # 查找该角色在当前行动空间的环境变量
                    role_vars = RoleVariable.query.filter_by(
                        role_id=agent.role_id,
                        action_space_id=new_task.action_space_id
                    ).all()

                    # 为智能体创建这些环境变量
                    for role_var in role_vars:
                        try:
                            # 将角色变量转换为智能体变量
                            AgentVariableService.create_variable(
                                agent_id=agent.id,
                                name=role_var.name,
                                value=role_var.value,
                                is_public=True,
                                label=role_var.label
                            )
                            logger.info(f"已为智能体 {agent.id} 创建环境变量: {role_var.name}")
                        except Exception as e:
                            logger.error(f"为智能体 {agent.id} 创建环境变量 {role_var.name} 失败: {str(e)}")

                # 检查是否已经有关联
                existing = ActionTaskAgent.query.filter_by(
                    action_task_id=new_task.id,
                    agent_id=agent_id
                ).first()

                if not existing:
                    # 创建行动任务-智能体关联
                    task_agent = ActionTaskAgent(
                        action_task_id=new_task.id,
                        agent_id=agent_id,
                        is_default=False  # 从行动空间角色创建的智能体已经设置了默认
                    )
                    db.session.add(task_agent)

        db.session.commit()

    # 创建默认会话
    from app.services.conversation_service import ConversationService
    default_conversation = ConversationService.create_conversation_for_action_task(new_task)

    # 获取所有智能体ID和信息
    task_agents = ActionTaskAgent.query.filter_by(action_task_id=new_task.id).all()
    agent_ids = []
    agent_info = []

    for ta in task_agents:
        agent = Agent.query.get(ta.agent_id)
        if agent:
            agent_ids.append(agent.id)

            # 获取角色信息
            role_name = None
            if hasattr(agent, 'role_id') and agent.role_id:
                role = Role.query.get(agent.role_id)
                if role:
                    role_name = role.name

            # 添加智能体信息
            agent_info.append({
                'id': agent.id,
                'name': agent.name,
                'role_name': role_name
            })

    # 初始化项目空间文件结构
    try:
        workspace_service.initialize_workspace_for_action_task(
            task_id=new_task.id,
            agent_ids=agent_ids,
            task_title=new_task.title,
            agent_info=agent_info
        )
        logger.info(f"已为行动任务 {new_task.id} 初始化项目空间文件结构")
    except Exception as e:
        logger.error(f"初始化项目空间文件结构失败: {str(e)}")
        # 继续处理，不中断任务创建

    # 使用任务详情的API获取结果
    task_detail = get_action_task(new_task.id)
    return task_detail

@router.put('/action-tasks/{task_id}')
async def update_action_task(task_id, request: Request, current_user=Depends(get_current_user)):
    """更新行动任务信息"""
    # 获取当前用户
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 检查编辑权限
    # ActionTask 使用 user_id 而不是 created_by
    if not current_user.is_admin and task.user_id != current_user.id:
        raise HTTPException(status_code=403, detail={'error': '无权限编辑此行动任务'})

    data = await request.json()

    # 更新基本信息
    if 'title' in data:
        task.title = data['title']
    if 'description' in data:
        task.description = data['description']
    if 'mode' in data:
        task.mode = data['mode']
    if 'status' in data:
        task.status = data['status']
    if 'rule_set_id' in data:
        task.rule_set_id = data['rule_set_id']
    if 'action_space_id' in data:
        task.action_space_id = data['action_space_id']

    # 只有创建者可以修改 is_shared 状态
    if 'is_shared' in data:
        if current_user.is_admin or task.user_id == current_user.id:
            task.is_shared = data['is_shared']

    db.session.commit()

    return {
        'id': task.id,
        'title': task.title,
        'message': '行动任务更新成功'
    }

@router.delete('/action-tasks/{task_id}')
def delete_action_task(task_id, request: Request, current_user=Depends(get_current_user)):
    """删除行动任务"""
    # 获取当前用户
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 检查删除权限
    # ActionTask 使用 user_id 而不是 created_by
    if not current_user.is_admin and task.user_id != current_user.id:
        raise HTTPException(status_code=403, detail={'error': '无权限删除此行动任务'})

    # 获取级联删除参数
    cascade = request.query_params.get('cascade', 'false').lower() == 'true'
    # 获取强制清理参数
    force_cleanup = request.query_params.get('force_cleanup', 'false').lower() == 'true'

    stopped_autonomous_tasks = 0

    try:
        if cascade:
            logger.info(f"开始级联删除行动任务 {task_id}，强制清理: {force_cleanup}")
            # 1. 停止并删除与任务关联的自主行动
            conversations = Conversation.query.filter_by(action_task_id=task_id).all()
            for conversation in conversations:
                # 获取该会话中的所有自主任务
                autonomous_tasks = AutonomousTask.query.filter_by(conversation_id=conversation.id).all()
                for autonomous_task in autonomous_tasks:
                    # 如果自主任务正在运行，尝试停止它
                    if autonomous_task.status == 'active':
                        try:
                            # 使用统一的stop_task接口停止任务
                            from app.services.scheduler.task_adapter import stop_task
                            stop_success = stop_task(task_id, conversation.id, autonomous_task.type)
                            if stop_success:
                                logger.info(f"已停止自主任务: {autonomous_task.id} (类型: {autonomous_task.type})")
                            else:
                                # 直接更新状态
                                autonomous_task.status = 'stopped'
                                logger.error(f"停止自主任务失败，已直接更新状态: {autonomous_task.id}")

                            stopped_autonomous_tasks += 1
                        except Exception as e:
                            logger.info(f"停止自主任务 {autonomous_task.id} 时出错: {str(e)}")
                            # 继续处理，不中断删除流程

                    # 删除自主任务的执行记录
                    AutonomousTaskExecution.query.filter_by(autonomous_task_id=autonomous_task.id).delete()

                # 删除该会话的所有自主任务
                AutonomousTask.query.filter_by(conversation_id=conversation.id).delete()

            logger.info(f"已停止并删除 {stopped_autonomous_tasks} 个自主任务")

            # 2. 删除与任务关联的会话及其数据
            for conversation in conversations:
                # 删除会话中的消息
                Message.query.filter_by(conversation_id=conversation.id).delete()
                # 删除会话中的智能体关联
                ConversationAgent.query.filter_by(conversation_id=conversation.id).delete()

            # 删除所有会话
            Conversation.query.filter_by(action_task_id=task_id).delete()

            # 2. 删除任务环境变量
            ActionTaskEnvironmentVariable.query.filter_by(action_task_id=task_id).delete()

            # 3. 删除与任务直接关联的智能体及其变量
            # 首先获取所有与任务关联的智能体
            task_agents = Agent.query.filter_by(action_task_id=task_id).all()
            for agent in task_agents:
                # 删除智能体的所有变量
                from app.models import AgentVariable
                deleted_vars = AgentVariable.query.filter_by(agent_id=agent.id).delete()
                logger.info(f"已删除智能体 {agent.id} 的 {deleted_vars} 个变量")

            # 提交变量删除
            db.session.flush()

            # 然后删除智能体本身
            Agent.query.filter_by(action_task_id=task_id).delete()

        # 4. 删除智能体与任务的关联关系
        ActionTaskAgent.query.filter_by(action_task_id=task_id).delete()

        # 5. 删除项目空间文件
        try:
            workspace_service.delete_workspace_for_action_task(task_id)
            logger.info(f"已删除行动任务 {task_id} 的项目空间文件")
        except Exception as e:
            logger.error(f"删除行动任务 {task_id} 的项目空间文件失败: {str(e)}")
            # 继续处理，不中断任务删除

        # 6. 如果启用强制清理，执行额外的清理操作
        if force_cleanup:
            logger.info("执行强制清理操作...")
            from sqlalchemy import text

            # 清理可能的孤立记录
            orphaned_executions = db.session.execute(text("""
                DELETE FROM autonomous_task_executions
                WHERE autonomous_task_id NOT IN (SELECT id FROM autonomous_tasks)
            """))
            logger.info(f"清理了 {orphaned_executions.rowcount} 条孤立的执行记录")

            orphaned_autonomous_tasks = db.session.execute(text("""
                DELETE FROM autonomous_tasks
                WHERE conversation_id NOT IN (SELECT id FROM conversations)
            """))
            logger.info(f"清理了 {orphaned_autonomous_tasks.rowcount} 条孤立的自主任务记录")

            orphaned_conv_agents = db.session.execute(text("""
                DELETE FROM conversation_agents
                WHERE conversation_id NOT IN (SELECT id FROM conversations)
                OR agent_id NOT IN (SELECT id FROM agents)
            """))
            logger.info(f"清理了 {orphaned_conv_agents.rowcount} 条孤立的会话智能体关联记录")

            orphaned_messages = db.session.execute(text("""
                DELETE FROM messages
                WHERE conversation_id IS NOT NULL
                AND conversation_id NOT IN (SELECT id FROM conversations)
            """))
            logger.info(f"清理了 {orphaned_messages.rowcount} 条孤立的消息记录")

            logger.info("强制清理完成")

        # 7. 最后删除行动任务本身
        db.session.delete(task)
        db.session.commit()

        logger.info(f"行动任务 {task_id} 删除完成")

        # 构建删除成功消息
        message_parts = []
        if stopped_autonomous_tasks > 0:
            message_parts.append(f'{stopped_autonomous_tasks} 个自主行动')
        message_parts.extend(['智能体', '环境变量', '会话记录', '长期记忆文件等'])

        return {
            'success': True,
            'message': f'行动任务及其所有关联数据已删除（包括{"、".join(message_parts)}）',
            'cascade': cascade,
            'force_cleanup': force_cleanup,
            'stopped_autonomous_tasks': stopped_autonomous_tasks
        }

    except Exception as e:
        db.session.rollback()
        logger.error(f"删除行动任务错误: {str(e)}")
        raise HTTPException(status_code=500, detail={
            'error': f'删除行动任务失败: {str(e)}',
            'details': str(e)
        })



# ============================================================
# Source: agents.py
# ============================================================

"""
Action Tasks Agent管理
"""

# (imports already at top of file)

@router.get('/action-tasks/{task_id}/agents')
def get_task_agents(task_id, request: Request):
    """获取行动任务的所有智能体"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 获取查询参数
    is_observer = request.query_params.get('is_observer')

    # 获取行动任务的智能体
    task_agents = ActionTaskAgent.query.filter_by(action_task_id=task_id).all()
    agents = []

    for ta in task_agents:
        agent = Agent.query.get(ta.agent_id)
        if agent:
            # 如果指定了is_observer参数，进行筛选
            if is_observer is not None:
                is_observer_bool = is_observer.lower() == 'true'
                if agent.is_observer != is_observer_bool:
                    continue

            # 获取智能体的角色信息
            agent_role = None
            if hasattr(agent, 'role_id') and agent.role_id:
                agent_role = Role.query.get(agent.role_id)

            role_name = agent_role.name if agent_role else "智能助手"

            agents.append({
                'id': agent.id,
                'name': agent.name,
                'role_name': role_name,
                'description': agent.description,
                'avatar': agent.avatar,
                'is_default': ta.is_default,
                'is_observer': agent.is_observer,  # 添加是否为监督者标记
            })

    return {'agents': agents}

@router.post('/action-tasks/{task_id}/agents')
async def add_agent_to_task(task_id, request: Request):
    """向行动任务添加智能体"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    data = await request.json()

    # 验证必填字段
    if not data.get('agent_id'):
        raise HTTPException(status_code=400, detail={'error': '缺少必填字段: agent_id'})

    agent_id = data.get('agent_id')
    agent = Agent.query.get(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail={'error': '智能体未找到'})

    # 检查智能体是否已经关联到任务
    existing = ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        agent_id=agent_id
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail={'error': '该智能体已经添加到任务中'})

    # 检查是否有其他智能体，决定是否设为默认
    has_other_agents = ActionTaskAgent.query.filter_by(action_task_id=task_id).count() > 0

    # 创建关联
    task_agent = ActionTaskAgent(
        action_task_id=task_id,
        agent_id=agent_id,
        is_default=not has_other_agents  # 如果没有其他智能体，则设为默认
    )

    db.session.add(task_agent)
    db.session.commit()

    return JSONResponse(content={
        'success': True,
        'message': '智能体已添加到任务',
        'agent': {
            'id': agent.id,
            'name': agent.name,
            'description': agent.description,
            'avatar': agent.avatar,
            'is_default': task_agent.is_default
        }
    }, status_code=201)

@router.delete('/action-tasks/{task_id}/agents/{agent_id}')
def remove_agent_from_task(task_id, agent_id):
    """从行动任务中移除智能体"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 查找关联记录
    task_agent = ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        agent_id=agent_id
    ).first()

    if not task_agent:
        raise HTTPException(status_code=404, detail={'error': '该智能体不在任务中'})

    # 如果是默认智能体且不是唯一的智能体，需要指定新的默认智能体
    if task_agent.is_default:
        other_agent = ActionTaskAgent.query.filter(
            ActionTaskAgent.action_task_id == task_id,
            ActionTaskAgent.agent_id != agent_id
        ).first()

        if other_agent:
            other_agent.is_default = True

    # 删除关联
    db.session.delete(task_agent)
    db.session.commit()

    return {
        'success': True,
        'message': '智能体已从任务中移除'
    }

@router.put('/action-tasks/{task_id}/agents/{agent_id}/default')
def set_default_agent(task_id, agent_id):
    """设置行动任务的默认智能体"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 确保智能体存在并且与任务关联
    task_agent = ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        agent_id=agent_id
    ).first()

    if not task_agent:
        raise HTTPException(status_code=404, detail={'error': '该智能体不在任务中'})

    # 清除其他默认智能体
    ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        is_default=True
    ).update({'is_default': False})

    # 设置新的默认智能体
    task_agent.is_default = True
    db.session.commit()

    return {
        'success': True,
        'message': '已设置默认智能体'
    }

@router.post('/action-tasks/{task_id}/agents/from-role')
async def add_agent_from_role(task_id, request: Request):
    """从角色创建智能体实例并添加到行动任务"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    data = await request.json()

    # 验证必填字段
    if not data.get('role_id'):
        raise HTTPException(status_code=400, detail={'error': '缺少必填字段: role_id'})

    role_id = data.get('role_id')
    role = Role.query.get(role_id)
    if not role:
        raise HTTPException(status_code=404, detail={'error': '角色未找到'})

    # 创建智能体实例（不继承环境变量和额外提示词）
    agent = Agent(
        name=data.get('name', role.name),
        description=data.get('description', role.description),
        avatar=data.get('avatar', role.avatar),
        settings=data.get('settings', role.settings),
        action_task_id=task_id,
        role_id=role_id
    )
    db.session.add(agent)
    db.session.flush()  # 获取agent.id

    # 检查是否有其他智能体，决定是否设为默认
    has_default_agent = ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        is_default=True
    ).first() is not None

    # 创建关联
    task_agent = ActionTaskAgent(
        action_task_id=task_id,
        agent_id=agent.id,
        is_default=not has_default_agent  # 如果没有默认智能体，则设为默认
    )

    db.session.add(task_agent)
    db.session.commit()

    # 为新智能体创建工作空间
    from app.models import SystemSetting
    if SystemSetting.get('create_agent_workspace', False):
        try:
            import os
            
            task_dir = os.path.join(workspace_service.workspace_dir, f'ActionTask-{task_id}')
            if os.path.exists(task_dir):
                # 创建智能体目录和工作空间文件
                agent_dir = os.path.join(task_dir, f'Agent-{agent.id}')
                os.makedirs(agent_dir, exist_ok=True)
                
                agent_display = f"{agent.name}[{role.name}][ID: {agent.id}]"
                workspace_service._create_agent_workspace_file(
                    os.path.join(agent_dir, 'AgentWorkspace.md'),
                    task_id, agent.id, task.title, agent_display
                )
                # 更新索引
                workspace_service.update_project_index_if_needed(task_id)
                logger.info(f"已为智能体 {agent.id} 创建工作空间")
        except Exception as e:
            logger.error(f"创建智能体工作空间失败: {e}")

    return JSONResponse(content={
        'success': True,
        'message': '智能体已添加到任务',
        'agent': {
            'id': agent.id,
            'name': agent.name,
            'description': agent.description,
            'avatar': agent.avatar,
            'role_id': role.id,
            'is_default': task_agent.is_default
        }
    }, status_code=201)

@router.put('/action-tasks/{task_id}/rule-set')
async def set_task_rule_set(task_id, request: Request):
    """设置行动任务的规则集"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    data = await request.json()
    rule_set_id = data.get('rule_set_id')

    if rule_set_id:
        # 确保规则集存在
        rule_set = RuleSet.query.get(rule_set_id)
        if not rule_set:
            raise HTTPException(status_code=404, detail={'error': '规则集未找到'})

    # 更新任务的规则集
    task.rule_set_id = rule_set_id
    db.session.commit()

    return {
        'success': True,
        'message': '已设置规则集',
        'rule_set_id': rule_set_id
    }



# ============================================================
# Source: environment.py
# ============================================================

"""
Action Tasks 环境变量管理
"""

# (imports already at top of file)

@router.get('/action-tasks/{task_id}/environment')
def get_task_environment(task_id, request: Request):
    """获取行动任务的环境变量"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 获取行动任务的环境变量
    variables = ActionTaskEnvironmentVariable.query.filter_by(action_task_id=task_id).all()
    result = []

    for var in variables:
        var_info = {
            'id': var.id,
            'name': var.name,  # 使用实际变量名
            'label': getattr(var, 'label', var.name.replace('_', ' ').title()),  # 优先使用label字段
            'value': var.value,
            'history': var.history if var.history else [],
            'is_readonly': getattr(var, 'is_readonly', False),  # 是否只读
            'source': 'shared' if getattr(var, 'shared_variable_id', None) else 'task'  # 标记来源
        }

        # 如果是共享环境变量，添加共享变量信息
        if getattr(var, 'shared_variable_id', None):
            var_info['shared_variable_id'] = var.shared_variable_id
            if hasattr(var, 'shared_variable') and var.shared_variable:
                var_info['shared_variable_name'] = var.shared_variable.name
                var_info['shared_variable_description'] = var.shared_variable.description

        result.append(var_info)

    # 获取任务相关的智能体变量
    # 先获取所有相关的智能体
    task_agents = ActionTaskAgent.query.filter_by(action_task_id=task_id).all()

    # 如果前端请求包含agent_variables参数，则也返回智能体变量
    if request.query_params.get('agent_variables', '').lower() == 'true':
        for task_agent in task_agents:
            agent = Agent.query.get(task_agent.agent_id)
            if agent:
                # 获取该智能体的变量
                agent_vars = AgentVariable.query.filter_by(agent_id=agent.id).all()
                for var in agent_vars:
                    result.append({
                        'id': var.id,
                        'name': var.name,
                        'label': getattr(var, 'label', var.name.replace('_', ' ').title()),  # 优先使用label字段
                        'value': var.value,
                        'is_public': var.is_public,
                        'agent_id': agent.id,
                        'agent_name': agent.name,
                        'role_id': agent.role_id,
                        'history': var.history if var.history else [],
                        'source': 'agent'
                    })

    return {'variables': result}

@router.put('/action-tasks/{task_id}/environment/variables')
async def update_task_environment_variable(task_id, request: Request):
    """更新行动任务的环境变量"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    data = await request.json()
    if not data.get('name'):
        raise HTTPException(status_code=400, detail={'error': '缺少必填字段: name'})

    # 使用实际变量名称（name字段）
    var_name = data.get('name')
    var_value = data.get('value')

    # 查找环境变量
    variable = ActionTaskEnvironmentVariable.query.filter_by(
        action_task_id=task_id,
        name=var_name
    ).first()

    if not variable:
        # 如果不存在，创建新的环境变量
        variable = ActionTaskEnvironmentVariable(
            name=var_name,
            value=var_value,
            type=data.get('type', 'text'),
            action_task_id=task_id
        )
        db.session.add(variable)
    else:
        # 检查是否为只读的共享环境变量
        if getattr(variable, 'is_readonly', False):
            raise HTTPException(status_code=400, detail={'error': '该环境变量为只读共享变量，不能修改'})

        # 更新历史记录
        # Ensure history is a list (may be string from DB or None)
        if not variable.history or not isinstance(variable.history, list):
            variable.history = []

        # 添加当前值到历史记录
        from app.utils.datetime_utils import get_current_time_with_timezone
        history_entry = {
            'value': variable.value,
            'timestamp': get_current_time_with_timezone().isoformat()
        }
        variable.history = variable.history + [history_entry]

        # 更新当前值
        variable.value = var_value

    # 更新行动任务的updated_at时间
    task.updated_at = get_current_time_with_timezone()

    db.session.commit()

    return {
        'id': variable.id,
        'name': variable.name,
        'label': variable.name.replace('_', ' ').title(),  # 生成一个显示标签
        'value': variable.value,
        'history': variable.history,
        'message': '环境变量更新成功'
    }

@router.delete('/action-tasks/{task_id}/environment/variables/{var_name}')
def delete_task_environment_variable(task_id, var_name):
    """删除行动任务的环境变量"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 查找环境变量
    variable = ActionTaskEnvironmentVariable.query.filter_by(
        action_task_id=task_id,
        name=var_name
    ).first()

    if not variable:
        raise HTTPException(status_code=404, detail={'error': f'环境变量 {var_name} 不存在'})

    # 检查是否为只读的共享环境变量
    if getattr(variable, 'is_readonly', False):
        raise HTTPException(status_code=400, detail={'error': '该环境变量为只读共享变量，不能删除'})

    # 删除变量
    db.session.delete(variable)

    # 更新行动任务的updated_at时间
    task.updated_at = get_current_time_with_timezone()

    db.session.commit()

    return JSONResponse(content={'message': f'环境变量 {var_name} 已删除'}, status_code=200)


@router.get('/action-tasks/{task_id}/agents/{agent_id}/variables')
def get_agent_variables(task_id, agent_id):
    """获取特定智能体的环境变量"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 验证智能体属于这个任务
    task_agent = ActionTaskAgent.query.filter_by(
        action_task_id=task_id,
        agent_id=agent_id
    ).first()

    if not task_agent:
        raise HTTPException(status_code=404, detail={'error': '智能体不属于该任务'})

    agent = Agent.query.get(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail={'error': '智能体未找到'})

    # 获取智能体的变量
    variables = AgentVariable.query.filter_by(agent_id=agent_id).all()
    result = []

    for var in variables:
        result.append({
            'id': var.id,
            'name': var.name,
            'label': getattr(var, 'label', var.name.replace('_', ' ').title()),  # 优先使用label字段
            'value': var.value,
            'is_public': var.is_public,
            'agent_id': agent_id,
            'agent_name': agent.name,
            'role_id': agent.role_id,
            'history': var.history if var.history else [],
            'source': 'agent'
        })

    return {
        'agent_id': agent_id,
        'agent_name': agent.name,
        'variables': result
    }



# ============================================================
# Source: export.py
# ============================================================

"""
Action Tasks 导出功能
"""

import tempfile
import zipfile
import shutil
from datetime import datetime

# (imports already at top of file)

@router.post('/action-tasks/{task_id}/export')
async def export_action_task(task_id, request: Request, current_user=Depends(get_current_user)):
    """导出行动任务数据"""
    try:
        # 验证任务是否存在和权限
        task = ActionTask.query.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

        # 检查用户权限
        if not can_access_task(task, current_user):
            raise HTTPException(status_code=403, detail={'error': '没有权限访问此任务'})

        # 获取请求参数
        data = await request.json() or {}
        include_agents = data.get('include_agents', True)
        conversations_scope = data.get('conversations_scope', 'all')  # 'all' or 'current'
        current_conversation_id = data.get('current_conversation_id')
        include_workspace = data.get('include_workspace', False)

        # 创建临时目录
        temp_dir = tempfile.mkdtemp()

        try:
            # 生成Excel文件
            excel_path = _create_export_excel(task, temp_dir, include_agents, conversations_scope, current_conversation_id)

            # 创建ZIP文件
            timestamp = datetime.now().strftime('%Y%m%d')
            zip_filename = f"{timestamp}-actiontask-{task.title}.zip"
            zip_path = os.path.join(temp_dir, zip_filename)

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 添加Excel文件
                zipf.write(excel_path, 'data.xlsx')

                # 如果需要包含工作空间
                if include_workspace:
                    _add_workspace_to_zip(task_id, zipf, temp_dir)

            # 读取ZIP文件内容到内存，然后清理临时目录
            import io
            with open(zip_path, 'rb') as f:
                zip_data = io.BytesIO(f.read())

            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)

            # 从内存返回文件
            zip_data.seek(0)
            return Response(
                content=zip_data.read(),
                media_type='application/zip',
                headers={'Content-Disposition': f'attachment; filename="{zip_filename}"'}
            )

        except Exception:
            # 确保异常情况下也清理临时文件
            shutil.rmtree(temp_dir, ignore_errors=True)
            raise

    except Exception as e:
        raise HTTPException(status_code=500, detail={'error': f'导出失败: {str(e)}'})

def _create_export_excel(task, temp_dir, include_agents, conversations_scope, current_conversation_id):
    """创建导出的Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Sheet1: 智能体列表
    if include_agents:
        ws_agents = wb.active
        ws_agents.title = "智能体列表"

        # 表头
        headers = ['ID', '名称', '角色', '描述', '状态', '创建时间', '是否默认', '是否监督者']
        for col, header in enumerate(headers, 1):
            cell = ws_agents.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 获取智能体数据
        task_agents = ActionTaskAgent.query.filter_by(action_task_id=task.id).all()
        row = 2
        for task_agent in task_agents:
            agent = Agent.query.get(task_agent.agent_id)
            if agent:
                role = Role.query.get(agent.role_id)
                ws_agents.cell(row=row, column=1, value=agent.id)
                ws_agents.cell(row=row, column=2, value=agent.name)
                ws_agents.cell(row=row, column=3, value=role.name if role else '')
                ws_agents.cell(row=row, column=4, value=agent.description or '')
                ws_agents.cell(row=row, column=5, value=agent.status)
                ws_agents.cell(row=row, column=6, value=agent.created_at.strftime('%Y-%m-%d %H:%M:%S') if agent.created_at else '')
                ws_agents.cell(row=row, column=7, value='是' if task_agent.is_default else '否')
                ws_agents.cell(row=row, column=8, value='是' if agent.is_observer else '否')
                row += 1

    # Sheet2: 会话消息
    ws_messages = wb.create_sheet("会话消息")

    # 表头
    msg_headers = ['时间', '会话', '角色', '发送者', '内容']
    for col, header in enumerate(msg_headers, 1):
        cell = ws_messages.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 获取消息数据
    if conversations_scope == 'current' and current_conversation_id:
        conversations = [Conversation.query.get(current_conversation_id)]
        conversations = [c for c in conversations if c and c.action_task_id == task.id]
    else:
        conversations = Conversation.query.filter_by(action_task_id=task.id).all()

    row = 2
    for conversation in conversations:
        messages = Message.query.filter_by(conversation_id=conversation.id).order_by(Message.created_at).all()
        for message in messages:
            # 获取发送者名称
            sender_name = ''
            if message.role == 'agent' and message.agent_id:
                agent = Agent.query.get(message.agent_id)
                sender_name = agent.name if agent else f'智能体{message.agent_id}'
            elif message.role == 'human':
                sender_name = '用户'
            elif message.role == 'system':
                sender_name = '系统'

            ws_messages.cell(row=row, column=1, value=message.created_at.strftime('%Y-%m-%d %H:%M:%S') if message.created_at else '')
            ws_messages.cell(row=row, column=2, value=conversation.title)
            ws_messages.cell(row=row, column=3, value=message.role)
            ws_messages.cell(row=row, column=4, value=sender_name)
            ws_messages.cell(row=row, column=5, value=message.content or '')
            row += 1

    # 保存文件
    excel_path = os.path.join(temp_dir, 'data.xlsx')
    wb.save(excel_path)
    return excel_path



def _add_workspace_to_zip(task_id, zipf, temp_dir):
    """将工作空间文件添加到ZIP中"""
    try:
        # 获取工作空间目录
        workspace_dir = workspace_service.workspace_dir
        task_workspace_dir = os.path.join(workspace_dir, f'ActionTask-{task_id}')

        if os.path.exists(task_workspace_dir):
            # 遍历工作空间目录，添加所有文件
            for root, dirs, files in os.walk(task_workspace_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # 计算相对路径
                    arcname = os.path.join('workspace', os.path.relpath(file_path, task_workspace_dir))
                    zipf.write(file_path, arcname)

    except Exception as e:
        logger.info(f"添加工作空间文件到ZIP时出错: {str(e)}")


# ============================================================
# Source: monitoring.py
# ============================================================

"""
Action Tasks 监控和观察者管理
"""

# (imports already at top of file)

@router.get('/action-tasks/{task_id}/observers')
def get_task_observers(task_id):
    """获取行动任务的所有监督者智能体"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    # 获取行动任务的监督者智能体
    observers = Agent.query.filter_by(action_task_id=task_id, is_observer=True).all()
    result = []

    for observer in observers:
        # 查找是否有关联
        task_agent = ActionTaskAgent.query.filter_by(
            action_task_id=task_id,
            agent_id=observer.id
        ).first()

        # 获取角色信息
        role = Role.query.get(observer.role_id) if observer.role_id else None

        result.append({
            'id': observer.id,
            'name': observer.name,
            'description': observer.description,
            'avatar': observer.avatar,
            'role_id': observer.role_id,
            'role_name': role.name if role else None,
            'type': observer.type,
            'is_observer': observer.is_observer,
            'additional_prompt': observer.additional_prompt
        })

    return {'observers': result}

@router.post('/action-tasks/{task_id}/observers')
async def add_observer_to_task(task_id, request: Request):
    """向行动任务添加监督者智能体实例"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

    data = await request.json()

    # 验证必填字段
    if not data.get('role_id'):
        raise HTTPException(status_code=400, detail={'error': '缺少必填字段: role_id'})

    role_id = data.get('role_id')
    role = Role.query.get(role_id)
    if not role:
        raise HTTPException(status_code=404, detail={'error': '角色未找到'})

    # 获取行动空间中监督者的额外提示词
    additional_prompt = ''
    if task.action_space_id:
        space_observer = ActionSpaceObserver.query.filter_by(
            action_space_id=task.action_space_id,
            role_id=role_id
        ).first()

        if space_observer:
            additional_prompt = space_observer.additional_prompt

    # 创建监督者智能体实例
    observer_agent = Agent(
        name=data.get('name', role.name),
        description=data.get('description', role.description),
        avatar=data.get('avatar', role.avatar),
        settings=data.get('settings', role.settings),
        action_task_id=task_id,
        role_id=role_id,
        type='observer',  # 标记为监督者类型
        is_observer=True,  # 标记为监督者
        additional_prompt=additional_prompt  # 设置额外提示词
    )
    db.session.add(observer_agent)
    db.session.flush()  # 获取observer_agent.id

    # 创建行动任务-智能体关联（监督者不设为默认智能体）
    task_agent = ActionTaskAgent(
        action_task_id=task_id,
        agent_id=observer_agent.id,
        is_default=False
    )
    db.session.add(task_agent)
    db.session.commit()

    return JSONResponse(content={
        'success': True,
        'message': '监督者已添加到任务',
        'observer': {
            'id': observer_agent.id,
            'name': observer_agent.name,
            'description': observer_agent.description,
            'avatar': observer_agent.avatar,
            'role_id': role.id,
            'type': observer_agent.type,
            'is_observer': observer_agent.is_observer,
            'additional_prompt': observer_agent.additional_prompt
        }
    }, status_code=201)

@router.get('/action-tasks/{task_id}/batch-variables')
def get_batch_variables(task_id):
    """批量获取任务的所有变量（环境变量和所有智能体变量）"""
    try:
        # 获取任务信息
        task = ActionTask.query.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail={'error': '行动任务未找到'})

        # 获取环境变量
        env_vars = ActionTaskEnvironmentVariable.query.filter_by(action_task_id=task_id).all()
        environment_variables = []

        for var in env_vars:
            environment_variables.append({
                'id': var.id,
                'name': var.name,
                'label': getattr(var, 'label', var.name.replace('_', ' ').title()),
                'value': var.value,
                'history': var.history if var.history else [],
                'source': 'task'
            })

        # 获取所有智能体
        task_agents = ActionTaskAgent.query.filter_by(action_task_id=task_id).all()
        agent_variables = []

        # 一次性获取所有智能体变量
        agent_ids = [ta.agent_id for ta in task_agents]
        agents_dict = {agent.id: agent for agent in Agent.query.filter(Agent.id.in_(agent_ids)).all()}

        # 批量查询所有智能体变量
        all_agent_vars = AgentVariable.query.filter(AgentVariable.agent_id.in_(agent_ids)).all()

        for var in all_agent_vars:
            agent = agents_dict.get(var.agent_id)
            if agent:
                agent_variables.append({
                    'id': var.id,
                    'name': var.name,
                    'label': getattr(var, 'label', var.name.replace('_', ' ').title()),
                    'value': var.value,
                    'is_public': var.is_public,
                    'agent_id': agent.id,
                    'agent_name': agent.name,
                    'role_id': agent.role_id if hasattr(agent, 'role_id') else None,
                    'history': var.history if var.history else [],
                    'source': 'agent'
                })

        # 返回结果
        return {
            'environment_variables': environment_variables,
            'agent_variables': agent_variables,
            'last_updated': datetime.utcnow().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail={'error': f'获取变量失败: {str(e)}'})



# ============================================================
# Source: rules.py
# ============================================================

"""
Action Tasks 规则和触发器管理
"""

# (imports already at top of file)

@router.get('/action-tasks/{task_id}/rules')
def get_task_rules(task_id):
    """获取任务关联的所有规则"""
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务不存在'})

    # 通过行动空间获取规则集，再获取规则
    action_space = task.action_space
    if not action_space:
        return {'rules': []}

    rules = []
    for rule_set_relation in action_space.rule_sets:
        rule_set = rule_set_relation.rule_set
        for rule_relation in rule_set.rules_relation:
            rule = rule_relation.rule
            rules.append({
                'id': rule.id,
                'name': rule.name,
                'description': rule.description,
                'content': rule.content,
                'category': rule.category,
                'type': rule.type,
                'is_active': rule.is_active,
                'rule_set_name': rule_set.name,
                'interpreter': rule.settings.get('interpreter', 'javascript') if rule.settings else 'javascript'
            })

    return {'rules': rules}

@router.get('/action-tasks/{task_id}/rule-variables')
def get_task_rule_variables(task_id):
    """获取任务的规则测试变量上下文

    返回所有可用于规则测试的变量，包括：
    1. 任务环境变量
    2. 智能体变量（公开的）
    3. 外部环境变量
    """
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务不存在'})

    try:
        variables = {}

        # 1. 获取任务环境变量
        task_env_vars = ActionTaskEnvironmentVariable.query.filter_by(action_task_id=task_id).all()
        for var in task_env_vars:
            variables[var.name] = var.value

        # 2. 获取智能体的公开变量
        from app.models import Agent, AgentVariable
        agents = Agent.query.filter_by(action_task_id=task_id).all()
        for agent in agents:
            agent_vars = AgentVariable.query.filter_by(
                agent_id=agent.id,
                is_public=True  # 只获取公开变量
            ).all()
            for var in agent_vars:
                # 使用 agent_name.variable_name 格式避免冲突
                var_key = f"{agent.name}.{var.name}"
                variables[var_key] = var.value

        # 3. 获取外部环境变量
        from app.models import ExternalEnvironmentVariable
        external_vars = ExternalEnvironmentVariable.query.filter_by(
            status='active'  # 只获取活跃的外部变量
        ).all()
        for var in external_vars:
            if var.value is not None:
                variables[var.name] = var.value

        return {
            'variables': variables,
            'variable_count': len(variables),
            'last_updated': datetime.utcnow().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail={'error': f'获取规则变量失败: {str(e)}'})

@router.get('/action-tasks/{task_id}/rule-triggers')
def get_task_rule_triggers(task_id, request: Request):
    """获取任务的规则触发记录

    支持分页和过滤参数：
    - page: 页码，默认1
    - per_page: 每页数量，默认20
    - rule_id: 过滤特定规则
    - trigger_type: 过滤触发类型 (manual, automatic, scheduled)
    - passed: 过滤通过状态 (true, false)
    """
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务不存在'})

    # 获取查询参数
    page = int(request.query_params.get('page', 1))
    per_page = min(int(request.query_params.get('per_page', 20)), 100)  # 限制最大每页数量
    rule_id_str = request.query_params.get('rule_id')
    rule_id = int(rule_id_str) if rule_id_str else None
    trigger_type = request.query_params.get('trigger_type')
    passed_filter = request.query_params.get('passed')

    try:
        # 构建查询
        query = RuleTriggerLog.query.filter_by(action_task_id=task_id)

        # 应用过滤条件
        if rule_id:
            query = query.filter(RuleTriggerLog.rule_id == rule_id)

        if trigger_type:
            query = query.filter(RuleTriggerLog.trigger_type == trigger_type)

        if passed_filter is not None:
            passed_bool = passed_filter.lower() == 'true'
            query = query.filter(RuleTriggerLog.passed == passed_bool)

        # 按创建时间倒序排列
        query = query.order_by(RuleTriggerLog.created_at.desc())

        # 手动分页（替代 Flask-SQLAlchemy 的 paginate）
        total = query.count()
        pages_count = (total + per_page - 1) // per_page
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 构建结果
        triggers = []
        for trigger in items:
            # 获取关联的规则信息
            rule = Rule.query.get(trigger.rule_id)

            trigger_data = {
                'id': trigger.id,
                'rule_id': trigger.rule_id,
                'rule_name': rule.name if rule else f'规则{trigger.rule_id}',
                'rule_type': rule.type if rule else 'unknown',
                'action_task_id': trigger.action_task_id,
                'conversation_id': trigger.conversation_id,
                'trigger_type': trigger.trigger_type,
                'trigger_source': trigger.trigger_source,
                'context': trigger.context,
                'variables': trigger.variables,
                'result': trigger.result,
                'passed': trigger.passed,
                'message': trigger.message,
                'details': trigger.details,
                'execution_time': trigger.execution_time,
                'created_at': trigger.created_at.isoformat() if trigger.created_at else None,
                'updated_at': trigger.updated_at.isoformat() if trigger.updated_at else None
            }
            triggers.append(trigger_data)

        return {
            'triggers': triggers,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': pages_count,
                'has_prev': page > 1,
                'has_next': page < pages_count
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail={'error': f'获取规则触发记录失败: {str(e)}'})

@router.post('/action-tasks/{task_id}/rule-triggers')
async def create_rule_trigger_log(task_id, request: Request):
    """创建规则触发记录

    请求体示例：
    {
        "rule_id": 1,
        "conversation_id": 1,
        "trigger_type": "manual",
        "trigger_source": "user",
        "context": "测试场景描述",
        "variables": {"var1": "value1"},
        "result": {"raw_result": true},
        "passed": true,
        "message": "规则检查通过",
        "details": "详细信息",
        "execution_time": 0.5
    }
    """
    task = ActionTask.query.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail={'error': '行动任务不存在'})
    data = await request.json()

    if not data:
        raise HTTPException(status_code=400, detail={'error': '缺少请求数据'})

    # 验证必填字段
    required_fields = ['rule_id']
    for field in required_fields:
        if field not in data:
            raise HTTPException(status_code=400, detail={'error': f'缺少必填字段: {field}'})

    try:
        # 验证规则是否存在
        rule = Rule.query.get(data['rule_id'])
        if not rule:
            raise HTTPException(status_code=404, detail={'error': f'规则ID {data["rule_id"]} 不存在'})

        # 创建触发记录
        trigger_log = RuleTriggerLog(
            rule_id=data['rule_id'],
            action_task_id=task_id,
            conversation_id=data.get('conversation_id'),
            trigger_type=data.get('trigger_type', 'manual'),
            trigger_source=data.get('trigger_source'),
            context=data.get('context'),
            variables=data.get('variables', {}),
            result=data.get('result'),
            passed=data.get('passed'),
            message=data.get('message'),
            details=data.get('details'),
            execution_time=data.get('execution_time')
        )

        db.session.add(trigger_log)
        db.session.commit()

        return JSONResponse(content={
            'id': trigger_log.id,
            'message': '规则触发记录创建成功'
        }, status_code=201)

    except Exception as e:
        db.session.rollback()
        raise HTTPException(status_code=500, detail={'error': f'创建规则触发记录失败: {str(e)}'})


